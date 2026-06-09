"""
Genera biohuerto_datos_ejemplo.xlsx con data realista para poblar
la app BioHuerto USAT y verificar que todos los modelos calzan.

Ejecutar:
  python datos_ejemplo/generar_datos.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────

COLOR_HEADER = "1B5E20"   # verde oscuro
COLOR_SUB    = "2E7D32"   # verde medio
COLOR_NOTA   = "FFF9C4"   # amarillo suave

def _hdr(ws, row, cols, color=COLOR_HEADER):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def _row(ws, row, vals):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def _autowidth(ws, min_w=12, max_w=40):
    for col in ws.columns:
        best = min_w
        for cell in col:
            try:
                best = max(best, min(max_w, len(str(cell.value or "")) + 4))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = best

def _nota(ws, row, text):
    cell = ws.cell(row=row, column=1, value=f"ℹ️  {text}")
    cell.fill = PatternFill("solid", fgColor=COLOR_NOTA)
    cell.font = Font(italic=True, size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column or 10)

# ── 1. HOJA: instrucciones ────────────────────────────────────────────────────
ws = wb.active
ws.title = "INSTRUCCIONES"
instrucciones = [
    ("ARCHIVO DE DATOS DE EJEMPLO - BioHuerto USAT", None),
    ("", None),
    ("Cada hoja corresponde a un modelo/tabla de la base de datos.", None),
    ("Sigue el orden indicado al cargar datos (respeta FKs).", None),
    ("", None),
    ("ORDEN DE CARGA", "HOJA"),
    ("1", "CATALOGO_Variedades"),
    ("2", "CATALOGO_Productos"),
    ("3", "CATALOGO_TiposLabor"),
    ("4", "CATALOGO_Plagas"),
    ("5", "CATALOGO_Objetivos"),
    ("6", "CATALOGO_Unidades"),
    ("7", "CATALOGO_Condiciones"),
    ("8", "Biohuerto"),
    ("9", "Campanas"),
    ("10", "Labores"),
    ("11", "PlanFitosanitario"),
    ("12", "Aplicaciones"),
    ("13", "PlanRiego"),
    ("14", "RegistrosRiego"),
    ("15", "Presupuesto"),
    ("16", "PracticasSostenibles"),
    ("17", "Cosechas"),
    ("18", "Pedidos"),
]
for r, (a, b) in enumerate(instrucciones, 1):
    ws.cell(r, 1, a).font = Font(bold=(r == 1 or a == "ORDEN DE CARGA"), size=11 if r == 1 else 10)
    if b:
        ws.cell(r, 2, b)
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 30

# ── 2. CATALOGO: Variedades ───────────────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Variedades")
_hdr(ws, 1, ["cultivo_macro", "nombre_variedad", "subtipo", "tipo_ciclo", "dias_ciclo", "descripcion"])
datos = [
    ("Lechuga",   "Batavia",            "Verde",       "anual",   60,  "Variedad de hoja suelta, resistente al calor, ideal para climas cálidos costeros."),
    ("Lechuga",   "Romana",             "Cos",         "anual",   70,  "Hoja alargada y crujiente. Mayor contenido de fibra que la batavia."),
    ("Cebolla",   "Roja Arequipeña",    "",            "anual",   120, "Bulbo rojo morado, sabor suave. Muy consumida en la costa norte del Perú."),
    ("Cebolla",   "Amarilla Dulce",     "",            "anual",   110, "Bulbo amarillo con sabor suave. Buena para exportación."),
    ("Tomate",    "Rio Grande",         "Determinate", "anual",   90,  "Fruto firme, ideal para consumo fresco y procesamiento. Planta compacta."),
    ("Albahaca",  "Genovesa",           "Dulce",       "anual",   45,  "Aromática muy demandada. Ciclo corto, múltiples cosechas escalonadas."),
    ("Culantro",  "Cimarrón",           "",            "anual",   40,  "Cilantro de hoja ancha. Resistente a plagas. Ciclo muy corto."),
    ("Rabanito",  "Cherry Belle",       "Rojo",        "anual",   28,  "Ciclo ultrarrápido, ideal para rotación intensiva."),
    ("Pimiento",  "Morrón Rojo",        "",            "anual",   95,  "Alta demanda en restaurantes. Requiere tutores y poda de formación."),
    ("Marigold",  "Africana",           "Tagetes",     "anual",   70,  "Planta compañera, repele nematodos. Se intercala con hortalizas."),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 3. CATALOGO: Productos agrícolas ─────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Productos")
_hdr(ws, 1, ["nombre", "tipo", "unidad", "precio_unitario_S/", "descripcion", "activo"])
datos = [
    ("Compost maduro",          "enmienda",       "kg",  1.50,  "Compost de residuos vegetales, 45 días maduración.",                "Sí"),
    ("Humus de lombriz",        "enmienda",       "kg",  3.00,  "Vermicompost de alta calidad. Aporta microorganismos benéficos.",   "Sí"),
    ("Biol fermentado",         "bioestimulante", "L",   2.50,  "Biol de estiércol de vacuno, 21 días fermentación.",               "Sí"),
    ("Trichoderma harzianum",   "biologico",      "kg",  18.00, "Hongo antagonista, controla Fusarium y Pythium.",                  "Sí"),
    ("Beauveria bassiana",      "biologico",      "kg",  22.00, "Hongo entomopatógeno, controla mosca blanca y trips.",             "Sí"),
    ("Bacillus subtilis",       "biologico",      "L",   25.00, "Bacteria antagonista de hongos foliares.",                        "Sí"),
    ("Sulfato de cobre",        "fitosanitario",  "kg",  8.50,  "Caldo bordelés casero. Control de mildiu y antracnosis.",          "Sí"),
    ("Jabón potásico",          "fitosanitario",  "L",   12.00, "Control de áfidos, cochinillas y araña roja. Bajo residuo.",       "Sí"),
    ("Nitrato de calcio",       "fertilizante",   "kg",  4.50,  "Fertirrigación. Previene tip-burn en lechuga.",                    "Sí"),
    ("Sulfato de potasio",      "fertilizante",   "kg",  5.20,  "Fase de engrosamiento/maduración. Mejora calidad de bulbos.",      "Sí"),
    ("Extracto de algas",       "bioestimulante", "L",   30.00, "Estimula enraizamiento y resistencia a estrés hídrico.",          "Sí"),
    ("Caldo sulfo-cálcico",     "fitosanitario",  "L",   6.00,  "Control preventivo de oídio y ácaros.",                           "Sí"),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 4. CATALOGO: Tipos de Labor ───────────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_TiposLabor")
_hdr(ws, 1, ["codigo", "nombre", "tipo", "unidad_default", "costo_unitario_default_S/", "activo"])
datos = [
    ("TL-0001", "Preparación de suelo (pala + rastrillo)", "otro",       "jornal", 45.00, "Sí"),
    ("TL-0002", "Elaboración de camas de siembra",          "formacion",  "jornal", 45.00, "Sí"),
    ("TL-0003", "Siembra directa",                          "otro",       "hora",   15.00, "Sí"),
    ("TL-0004", "Trasplante de plántulas",                  "formacion",  "hora",   15.00, "Sí"),
    ("TL-0005", "Riego manual",                             "riego",      "hora",   12.00, "Sí"),
    ("TL-0006", "Aplicación foliar manual",                 "aplicacion", "hora",   15.00, "Sí"),
    ("TL-0007", "Deshierbe / escarda",                      "otro",       "jornal", 40.00, "Sí"),
    ("TL-0008", "Poda de formación",                        "poda",       "hora",   15.00, "Sí"),
    ("TL-0009", "Tutoraje",                                 "formacion",  "hora",   15.00, "Sí"),
    ("TL-0010", "Cosecha manual",                           "cosecha",    "jornal", 45.00, "Sí"),
    ("TL-0011", "Clasificación y empaque",                  "cosecha",    "hora",   12.00, "Sí"),
    ("TL-0012", "Fertirriego (goteo)",                      "riego",      "hora",   10.00, "Sí"),
    ("TL-0013", "Monitoreo fitosanitario",                  "otro",       "hora",   15.00, "Sí"),
    ("TL-0014", "Elaboración de compost in situ",           "otro",       "jornal", 45.00, "Sí"),
    ("TL-0015", "Instalación de cinta de goteo",            "otro",       "jornal", 45.00, "Sí"),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 5. CATALOGO: Plagas ───────────────────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Plagas")
_hdr(ws, 1, ["nombre", "nombre_cientifico", "tipo", "descripcion"])
datos = [
    ("Mosca blanca",    "Bemisia tabaci",         "insecto",  "Succionador de savia. Vector de virus. Alta presencia en verano costeño."),
    ("Pulgón verde",    "Myzus persicae",          "insecto",  "Colonia en brotes y envés de hojas. Exuda melaza que promueve fumagina."),
    ("Trips del tomate","Frankliniella occidentalis","insecto", "Daño foliar y en frutos. Vector de TSWV."),
    ("Araña roja",      "Tetranychus urticae",     "acaro",    "Punteado clorótico en hojas. Alta presión en épocas secas."),
    ("Fusarium",        "Fusarium oxysporum",      "hongo",    "Marchitez vascular. Suelo infestado. No tiene cura, solo prevención."),
    ("Mildiu velloso",  "Bremia lactucae",         "hongo",    "Manchas amarillas en haz y micelio gris en envés. Favorecido por humedad alta."),
    ("Oídio",           "Leveillula taurica",      "hongo",    "Polvillo blanco en haz. Frecuente en pimiento bajo condiciones secas."),
    ("Nematodo agallador","Meloidogyne incognita",  "nematodo", "Agallas en raíces. Reduce absorción de nutrientes y agua."),
    ("Gusano de tierra","Agrotis ipsilon",         "insecto",  "Larva nocturna. Corta tallos a nivel del suelo en etapa de plántula."),
    ("Pudrición blanda","Pectobacterium carotovorum","bacteria","Tejido acuoso y maloliente. Favorecida por exceso de humedad."),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 6. CATALOGO: Objetivos ────────────────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Objetivos")
_hdr(ws, 1, ["nombre", "tipo", "plaga_relacionada"])
datos = [
    ("Control de mosca blanca",        "control",       "Mosca blanca"),
    ("Control de pulgones",            "control",       "Pulgón verde"),
    ("Control de trips",               "control",       "Trips del tomate"),
    ("Control de araña roja",          "control",       "Araña roja"),
    ("Prevención de Fusarium",         "prevencion",    "Fusarium"),
    ("Control de mildiu",              "control",       "Mildiu velloso"),
    ("Control de oídio",               "control",       "Oídio"),
    ("Prevención nematodos",           "prevencion",    "Nematodo agallador"),
    ("Fertilización nitrogenada",      "fertilizacion", ""),
    ("Fertilización potásica",         "fertilizacion", ""),
    ("Bioestimulación de raíces",      "estimulacion",  ""),
    ("Bioestimulación foliar",         "estimulacion",  ""),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 7. CATALOGO: Unidades de medida ──────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Unidades")
_hdr(ws, 1, ["codigo", "nombre", "tipo"])
datos = [
    ("L/ha",   "Litros por hectárea",          "vol_area"),
    ("mL/ha",  "Mililitros por hectárea",       "vol_area"),
    ("L/m²",   "Litros por metro cuadrado",     "vol_area"),
    ("mL/m²",  "Mililitros por metro cuadrado", "vol_area"),
    ("kg/ha",  "Kilogramos por hectárea",       "masa_area"),
    ("g/ha",   "Gramos por hectárea",           "masa_area"),
    ("kg/m²",  "Kilogramos por metro cuadrado", "masa_area"),
    ("g/m²",   "Gramos por metro cuadrado",     "masa_area"),
    ("mL/L",   "Mililitros por litro de agua",  "vol_vol"),
    ("g/L",    "Gramos por litro de agua",      "masa_vol"),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 8. CATALOGO: Condiciones ──────────────────────────────────────────────────
ws = wb.create_sheet("CATALOGO_Condiciones")
_hdr(ws, 1, ["nombre", "descripcion"])
datos = [
    ("Alta humedad relativa (>75%)",  "Aplicar preventivos cuando HR supere 75% por ≥3 días consecutivos."),
    ("Temperatura >28°C",             "Condición de estrés térmico. Aplicar bioestimulantes."),
    ("Presencia visual de plaga",     "Al detectar 5% de plantas afectadas."),
    ("Al inicio de cada etapa",       "Aplicar al cambio de etapa fenológica."),
    ("Cada 7 días",                   "Aplicación preventiva semanal."),
    ("Cada 15 días",                  "Aplicación preventiva quincenal."),
    ("Post-lluvia o riego excesivo",  "Aplicar fungicida preventivo 24h después de lluvia o anegamiento."),
    ("Al trasplante",                 "Aplicar al momento del trasplante para favorecer enraizamiento."),
]
for i, d in enumerate(datos, 2):
    _row(ws, i, d)
_autowidth(ws)

# ── 9. BIOHUERTO ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Biohuerto")
_hdr(ws, 1, ["nombre", "codigo_auto", "area_m2", "ubicacion_referencia", "descripcion",
              "departamento", "provincia", "distrito", "latitud", "longitud", "activo"])
_nota(ws, 2, "El codigo (BH-001...) lo asigna el sistema. La lat/lng se obtiene del mapa al registrar.")
_row(ws, 3, [
    "Biohuerto USAT – Bloque Norte",
    "BH-001",
    480.00,
    "Campus USAT, pabellón de Ingeniería, costado del vivero institucional",
    "Biohuerto experimental destinado a la producción orgánica de hortalizas, "
    "aromáticas y plantas compañeras. Cuenta con sistema de riego por goteo en "
    "tres sectores, composteras in situ y área de semillero bajo cobertura.",
    "Lambayeque",
    "Chiclayo",
    "Chiclayo",
    -6.7741,
    -79.8455,
    "Sí",
])
_autowidth(ws)

# ── 10. CAMPAÑAS ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Campanas")
_hdr(ws, 1, ["#", "biohuerto", "variedad_cultivo", "variedad_nombre", "codigo",
              "anio", "fecha_inicio", "fecha_fin", "area_m2", "unidad_area",
              "estado", "tipo_ciclo", "objetivo_cosecha", "unidad_cosecha",
              "precio_venta_estimado_S/", "notas"])
_nota(ws, 2, "Campaña 1 = cerrada (historial). Campaña 2 = activa (en curso). Campaña 3 = planificada.")
campanas = [
    (1, "Biohuerto USAT – Bloque Norte", "Lechuga", "Batavia", "CAM-BH001-2025-001",
     2025, "2025-02-10", "2025-04-15", 50.00, "m²", "cerrada", "anual",
     75.00, "kg", 3.50,
     "Primera campaña del año. Ciclo exitoso, 78 kg cosechados. Problema leve de mosca blanca en semana 6."),
    (2, "Biohuerto USAT – Bloque Norte", "Cebolla", "Roja Arequipeña", "CAM-BH001-2025-002",
     2025, "2025-04-01", "2025-07-31", 80.00, "m²", "activa", "anual",
     160.00, "kg", 2.80,
     "Campaña en curso. Buen desarrollo vegetativo. Aplicación de Trichoderma al trasplante."),
    (3, "Biohuerto USAT – Bloque Norte", "Tomate", "Rio Grande", "CAM-BH001-2025-003",
     2025, "2025-08-01", "2025-10-31", 60.00, "m²", "planificada", "anual",
     120.00, "kg", 4.50,
     "Planificada para el segundo semestre. Requiere instalación de tutores y malla anti-insectos."),
]
for c in campanas:
    _row(ws, c[0] + 2, c)
_autowidth(ws)

# ── 11. LABORES ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Labores")
_hdr(ws, 1, ["#", "campana_codigo", "tipo_labor", "descripcion", "etapa",
              "fecha_programada", "fecha_ejecutada", "cantidad_programada",
              "cantidad_ejecutada", "unidad", "costo_unitario_S/", "operario",
              "estado", "notas"])
_nota(ws, 2, "Labores de CAM-BH001-2025-001 (cerrada) y CAM-BH001-2025-002 (activa).")
labores = [
    # Campaña 1 – Lechuga (cerrada)
    (1,  "CAM-BH001-2025-001", "Preparación de suelo (pala + rastrillo)", "",                      "preparacion", "2025-02-10", "2025-02-10", 1.0, 1.0, "jornal", 45.00, "Carlos R.",  "ejecutada", "Se incorporó 20 kg de compost maduro."),
    (2,  "CAM-BH001-2025-001", "Elaboración de camas de siembra",         "4 camas de 1×10 m",    "preparacion", "2025-02-11", "2025-02-11", 0.5, 0.5, "jornal", 45.00, "Carlos R.",  "ejecutada", "Camas elevadas 20 cm, bordes de madera reciclada."),
    (3,  "CAM-BH001-2025-001", "Siembra directa",                         "Almácigo en bandejas", "germinacion", "2025-02-12", "2025-02-12", 2.0, 2.0, "hora",   15.00, "Ana M.",     "ejecutada", "50 bandejas de 72 celdas. Sustrato: compost+arena 2:1."),
    (4,  "CAM-BH001-2025-001", "Trasplante de plántulas",                 "A los 20 días",        "crecimiento", "2025-03-04", "2025-03-04", 3.0, 3.0, "hora",   15.00, "Ana M.",     "ejecutada", "Marco 0.25×0.25 m. 800 plantas trasplantadas."),
    (5,  "CAM-BH001-2025-001", "Deshierbe / escarda",                     "1ra escarda",          "crecimiento", "2025-03-18", "2025-03-18", 0.5, 0.5, "jornal", 40.00, "Carlos R.",  "ejecutada", "Maleza principalmente oxalis y enredadera."),
    (6,  "CAM-BH001-2025-001", "Monitoreo fitosanitario",                 "Revisión semanal",     "floracion",   "2025-03-25", "2025-03-25", 1.0, 1.0, "hora",   15.00, "Prof. López","ejecutada", "Detección de mosca blanca en 8% de plantas. Aplicación inmediata."),
    (7,  "CAM-BH001-2025-001", "Cosecha manual",                         "Cosecha escalonada",   "cosecha",     "2025-04-08", "2025-04-08", 1.0, 1.0, "jornal", 45.00, "Carlos R.",  "ejecutada", "Primera pasada: 42 kg. Segunda (15 abr): 36 kg."),
    (8,  "CAM-BH001-2025-001", "Clasificación y empaque",                "",                      "cosecha",     "2025-04-08", "2025-04-08", 2.0, 2.0, "hora",   12.00, "Ana M.",     "ejecutada", "Empaque en bolsas de 250g. Total: 78 kg."),
    # Campaña 2 – Cebolla (activa)
    (9,  "CAM-BH001-2025-002", "Preparación de suelo (pala + rastrillo)", "Surcado profundo",     "preparacion", "2025-04-01", "2025-04-01", 1.0, 1.0, "jornal", 45.00, "Carlos R.",  "ejecutada", "Surcado a 30 cm. Incorporación de 30 kg humus de lombriz."),
    (10, "CAM-BH001-2025-002", "Siembra directa",                         "Siembra en surcos",    "germinacion", "2025-04-05", "2025-04-05", 2.0, 2.0, "hora",   15.00, "Ana M.",     "ejecutada", "3 semillas por golpe, cada 10 cm. 1.2 kg de semilla."),
    (11, "CAM-BH001-2025-002", "Deshierbe / escarda",                     "1ra escarda",          "crecimiento", "2025-04-28", "2025-04-28", 0.5, 0.5, "jornal", 40.00, "Carlos R.",  "ejecutada", "Raleo simultaneo: 1 planta por golpe."),
    (12, "CAM-BH001-2025-002", "Monitoreo fitosanitario",                 "Revisión semanal",     "crecimiento", "2025-05-15", "2025-05-15", 1.0, 1.0, "hora",   15.00, "Prof. López","ejecutada", "Sin incidencias. Desarrollo uniforme."),
    (13, "CAM-BH001-2025-002", "Aplicación foliar manual",               "Biol + sulfato potasio","crecimiento", "2025-05-20", "2025-05-20", 1.5, 1.5, "hora",   15.00, "Ana M.",     "ejecutada", "Aplicación foliar de biol al 20%. 15 L totales."),
    (14, "CAM-BH001-2025-002", "Deshierbe / escarda",                     "2da escarda",          "floracion",   "2025-06-10", None,         0.5, None, "jornal", 40.00, "Carlos R.",  "programada","Aporque ligero post escarda."),
    (15, "CAM-BH001-2025-002", "Cosecha manual",                         "",                      "cosecha",     "2025-07-20", None,         2.0, None, "jornal", 45.00, "Carlos R.",  "programada","Estimado 160 kg. Cosechar cuando cuello esté seco 50%."),
]
for l in labores:
    _row(ws, l[0] + 2, l)
_autowidth(ws)

# ── 12. PLAN FITOSANITARIO ────────────────────────────────────────────────────
ws = wb.create_sheet("PlanFitosanitario")
_hdr(ws, 1, ["#", "campana_codigo", "producto", "objetivo", "plaga",
              "etapa", "dosis", "unidad_dosis", "dias_antes_cosecha",
              "frecuencia_dias", "condicion", "estado", "fecha_aplicada", "activo"])
plan = [
    # Campaña 1 (cerrada)
    (1,  "CAM-BH001-2025-001", "Trichoderma harzianum",  "Prevención de Fusarium",   "Fusarium",      "preparacion", 2.50,  "kg/m²",  None, None, "Al inicio de cada etapa",  "aplicado", "2025-02-10", "Sí"),
    (2,  "CAM-BH001-2025-001", "Beauveria bassiana",     "Control de mosca blanca",  "Mosca blanca",  "crecimiento", 2.00,  "g/L",    7,    14,   "Presencia visual de plaga","aplicado", "2025-03-26", "Sí"),
    (3,  "CAM-BH001-2025-001", "Jabón potásico",         "Control de pulgones",      "Pulgón verde",  "crecimiento", 5.00,  "mL/L",   3,    7,    "Presencia visual de plaga","aplicado", "2025-03-26", "Sí"),
    (4,  "CAM-BH001-2025-001", "Biol fermentado",        "Fertilización nitrogenada","",              "germinacion", 10.00, "mL/L",   None, 14,   "Cada 15 días",             "aplicado", "2025-02-25", "Sí"),
    # Campaña 2 (activa)
    (5,  "CAM-BH001-2025-002", "Trichoderma harzianum",  "Prevención de Fusarium",   "Fusarium",      "preparacion", 2.50,  "kg/m²",  None, None, "Al inicio de cada etapa",  "aplicado", "2025-04-01", "Sí"),
    (6,  "CAM-BH001-2025-002", "Biol fermentado",        "Bioestimulación foliar",   "",              "crecimiento", 10.00, "mL/L",   None, 14,   "Cada 15 días",             "aplicado", "2025-05-20", "Sí"),
    (7,  "CAM-BH001-2025-002", "Sulfato de potasio",     "Fertilización potásica",   "",              "floracion",   5.00,  "g/m²",   None, None, "Al inicio de cada etapa",  "programado",None,        "Sí"),
    (8,  "CAM-BH001-2025-002", "Caldo sulfo-cálcico",    "Control de oídio",         "Oídio",         "floracion",   3.00,  "mL/L",   5,    10,   "Alta humedad relativa (>75%)", "programado", None,     "Sí"),
]
for p in plan:
    _row(ws, p[0] + 2, p)
_autowidth(ws)

# ── 13. APLICACIONES (ejecutadas) ─────────────────────────────────────────────
ws = wb.create_sheet("Aplicaciones")
_hdr(ws, 1, ["#", "campana_codigo", "item_plan_#", "producto",
              "fecha", "area_aplicada_m2", "dosis_aplicada",
              "costo_total_S/", "operario", "es_sostenible", "observaciones"])
aplicaciones = [
    (1, "CAM-BH001-2025-001", 1, "Trichoderma harzianum",  "2025-02-10", 50.00, 125.00,  56.25,  "Carlos R.",  "Sí", "Incorporado al suelo al preparar las camas."),
    (2, "CAM-BH001-2025-001", 4, "Biol fermentado",        "2025-02-25", 50.00, 0.010,    3.75,  "Ana M.",    "Sí", "Dilución 1:10. Aplicación foliar con mochila 15L."),
    (3, "CAM-BH001-2025-001", 4, "Biol fermentado",        "2025-03-11", 50.00, 0.010,    3.75,  "Ana M.",    "Sí", "2da aplicación. Buen color foliar observado."),
    (4, "CAM-BH001-2025-001", 2, "Beauveria bassiana",     "2025-03-26", 50.00, 100.00,  55.00,  "Carlos R.", "Sí", "Respuesta a detección de mosca blanca. Aplicar al atardecer."),
    (5, "CAM-BH001-2025-001", 3, "Jabón potásico",         "2025-03-26", 50.00, 5.00,    3.00,  "Carlos R.", "Sí", "Aplicado simultáneamente con Beauveria. 10L de mezcla."),
    (6, "CAM-BH001-2025-002", 5, "Trichoderma harzianum",  "2025-04-01", 80.00, 200.00,  90.00,  "Carlos R.", "Sí", "Incorporado al suelo antes de siembra."),
    (7, "CAM-BH001-2025-002", 6, "Biol fermentado",        "2025-05-20", 80.00, 0.010,    6.00,  "Ana M.",   "Sí", "Foliar. Aplicar en horas de menor temperatura."),
]
for a in aplicaciones:
    _row(ws, a[0] + 2, a)
_autowidth(ws)

# ── 14. PLAN RIEGO ────────────────────────────────────────────────────────────
ws = wb.create_sheet("PlanRiego")
_hdr(ws, 1, ["#", "campana_codigo", "nombre", "metodo", "litros_por_m2",
              "frecuencia_dias", "duracion_minutos", "fertilizante",
              "dosis_fertilizante_kg_m2", "fecha_inicio", "fecha_fin", "activo"])
riegos = [
    # Campaña 1
    (1, "CAM-BH001-2025-001", "Riego germinación – diario",    "goteo", 2.00, 1,  20, "",                   None,  "2025-02-12", "2025-03-03", "Sí"),
    (2, "CAM-BH001-2025-001", "Riego crecimiento – día por medio", "goteo", 3.00, 2, 30, "Nitrato de calcio", 0.003, "2025-03-04", "2025-04-07", "Sí"),
    # Campaña 2
    (3, "CAM-BH001-2025-002", "Riego inicial – día por medio", "goteo", 2.50, 2,  25, "",                   None,  "2025-04-05", "2025-05-04", "Sí"),
    (4, "CAM-BH001-2025-002", "Riego bulbificación – fertirriego", "goteo", 3.50, 2, 35, "Sulfato de potasio", 0.005, "2025-05-05", "2025-07-19", "Sí"),
]
for r in riegos:
    _row(ws, r[0] + 2, r)
_autowidth(ws)

# ── 15. REGISTROS DE RIEGO ────────────────────────────────────────────────────
ws = wb.create_sheet("RegistrosRiego")
_hdr(ws, 1, ["#", "campana_codigo", "plan_nombre", "fecha", "area_regada_m2",
              "litros_aplicados", "costo_agua_S/", "fertilizante",
              "dosis_fertilizante", "operario", "observaciones"])
reg_riego = [
    (1,  "CAM-BH001-2025-001", "Riego germinación – diario",        "2025-02-13", 50, 100.00, 0.80, "",                   None,  "Carlos R.", "Presión estable. Sin fugas."),
    (2,  "CAM-BH001-2025-001", "Riego germinación – diario",        "2025-02-20", 50, 100.00, 0.80, "",                   None,  "Carlos R.", ""),
    (3,  "CAM-BH001-2025-001", "Riego crecimiento – día por medio", "2025-03-06", 50, 150.00, 1.20, "Nitrato de calcio",  0.003, "Ana M.",    "Primera fertirriego. Disolución previa en bidón."),
    (4,  "CAM-BH001-2025-001", "Riego crecimiento – día por medio", "2025-03-20", 50, 150.00, 1.20, "Nitrato de calcio",  0.003, "Ana M.",    ""),
    (5,  "CAM-BH001-2025-001", "Riego crecimiento – día por medio", "2025-04-03", 50, 150.00, 1.20, "Nitrato de calcio",  0.003, "Ana M.",    "Último riego antes de cosecha."),
    (6,  "CAM-BH001-2025-002", "Riego inicial – día por medio",     "2025-04-07", 80, 200.00, 1.60, "",                   None,  "Carlos R.", "Germinación pareja observada a los 6 días."),
    (7,  "CAM-BH001-2025-002", "Riego inicial – día por medio",     "2025-04-21", 80, 200.00, 1.60, "",                   None,  "Carlos R.", ""),
    (8,  "CAM-BH001-2025-002", "Riego bulbificación – fertirriego", "2025-05-07", 80, 280.00, 2.24, "Sulfato de potasio", 0.005, "Ana M.",    "Inicio de fertirriego potásico."),
    (9,  "CAM-BH001-2025-002", "Riego bulbificación – fertirriego", "2025-05-21", 80, 280.00, 2.24, "Sulfato de potasio", 0.005, "Ana M.",    ""),
    (10, "CAM-BH001-2025-002", "Riego bulbificación – fertirriego", "2025-06-04", 80, 280.00, 2.24, "Sulfato de potasio", 0.005, "Ana M.",    ""),
]
for r in reg_riego:
    _row(ws, r[0] + 2, r)
_autowidth(ws)

# ── 16. PRESUPUESTO ───────────────────────────────────────────────────────────
ws = wb.create_sheet("Presupuesto")
_hdr(ws, 1, ["#", "campana_codigo", "categoria", "descripcion",
              "cantidad", "unidad", "precio_unitario_S/",
              "monto_presupuestado_S/", "monto_ejecutado_S/", "varianza_S/"])
presupuesto = [
    # Campaña 1 – Lechuga (cerrada – ejecutado completo)
    (1,  "CAM-BH001-2025-001", "insumo",    "Trichoderma harzianum 125 g",    1.0,  "sobre",  56.25,  56.25,  56.25,  0.00),
    (2,  "CAM-BH001-2025-001", "insumo",    "Biol fermentado 5 L",            2.0,  "balde",  12.50,  25.00,  22.50,  2.50),
    (3,  "CAM-BH001-2025-001", "insumo",    "Beauveria bassiana 100 g",       1.0,  "sobre",  55.00,  55.00,  55.00,  0.00),
    (4,  "CAM-BH001-2025-001", "insumo",    "Jabón potásico 1 L",             1.0,  "unid.",  12.00,  12.00,  12.00,  0.00),
    (5,  "CAM-BH001-2025-001", "insumo",    "Nitrato de calcio 0.45 kg",      3.0,  "sobre",   2.03,   6.08,   6.08,  0.00),
    (6,  "CAM-BH001-2025-001", "insumo",    "Compost maduro 20 kg",          20.0,  "kg",      1.50,  30.00,  28.00,  2.00),
    (7,  "CAM-BH001-2025-001", "agua",      "Consumo agua de riego (m³)",     2.5,  "m³",      1.00,   2.50,   2.48,  0.02),
    (8,  "CAM-BH001-2025-001", "mano_obra", "Mano de obra (jornales)",        4.0,  "jornal", 45.00, 180.00, 180.00,  0.00),
    (9,  "CAM-BH001-2025-001", "mano_obra", "Mano de obra (horas sueltas)",  10.0,  "hora",   12.00, 120.00, 108.00, 12.00),
    (10, "CAM-BH001-2025-001", "otro",      "Bolsas empaque 250g (x200)",     2.0,  "ciento",  8.00,  16.00,  16.00,  0.00),
    # Campaña 2 – Cebolla (activa – ejecutado parcial)
    (11, "CAM-BH001-2025-002", "insumo",    "Trichoderma harzianum 200 g",    1.0,  "sobre",  90.00,  90.00,  90.00,  0.00),
    (12, "CAM-BH001-2025-002", "insumo",    "Humus de lombriz 30 kg",        30.0,  "kg",      3.00,  90.00,  87.00,  3.00),
    (13, "CAM-BH001-2025-002", "insumo",    "Biol fermentado 10 L",           2.0,  "balde",  12.50,  25.00,  12.50, 12.50),
    (14, "CAM-BH001-2025-002", "insumo",    "Sulfato de potasio 2 kg",        2.0,  "kg",      5.20,  10.40,   0.00, 10.40),
    (15, "CAM-BH001-2025-002", "insumo",    "Caldo sulfo-cálcico 1 L",        1.0,  "L",       6.00,   6.00,   0.00,  6.00),
    (16, "CAM-BH001-2025-002", "insumo",    "Semilla cebolla roja 1.2 kg",    1.0,  "bolsa",  35.00,  35.00,  35.00,  0.00),
    (17, "CAM-BH001-2025-002", "agua",      "Consumo agua de riego (m³)",     6.0,  "m³",      1.00,   6.00,   2.96,  3.04),
    (18, "CAM-BH001-2025-002", "mano_obra", "Mano de obra (jornales)",        8.0,  "jornal", 45.00, 360.00, 135.00,225.00),
    (19, "CAM-BH001-2025-002", "mano_obra", "Mano de obra (horas sueltas)",  12.0,  "hora",   15.00, 180.00,  67.50,112.50),
    (20, "CAM-BH001-2025-002", "otro",      "Malla raschel 50% (2×10 m)",     1.0,  "rollo",  45.00,  45.00,  45.00,  0.00),
]
for p in presupuesto:
    _row(ws, p[0] + 2, p)
_autowidth(ws)

# ── 17. PRÁCTICAS SOSTENIBLES ─────────────────────────────────────────────────
ws = wb.create_sheet("PracticasSostenibles")
_hdr(ws, 1, ["#", "campana_codigo", "tipo", "fecha", "descripcion", "cantidad", "unidad"])
practicas = [
    (1, "CAM-BH001-2025-001", "compost",           "2025-02-10", "Incorporación de compost maduro (45 días) elaborado en la compostera del biohuerto con residuos del comedor USAT.", 20.00, "kg"),
    (2, "CAM-BH001-2025-001", "control_biologico", "2025-03-26", "Aplicación de Beauveria bassiana y jabón potásico para control de mosca blanca. Sin uso de insecticidas sintéticos.", 1.00, "jornal"),
    (3, "CAM-BH001-2025-001", "sin_agroquimicos",  "2025-04-15", "Campaña completada sin uso de agroquímicos sintéticos. 100% manejo orgánico certificado.", None, ""),
    (4, "CAM-BH001-2025-002", "compost",           "2025-04-01", "Incorporación de humus de lombriz de vermicompostera institucional.", 30.00, "kg"),
    (5, "CAM-BH001-2025-002", "riego_eficiente",   "2025-04-05", "Sistema de goteo instalado con cinta DripNet 16 mm, 30 cm entre goteros. Eficiencia estimada 92%.", 80.00, "m²"),
    (6, "CAM-BH001-2025-002", "control_biologico", "2025-04-01", "Inoculación con Trichoderma al suelo antes de siembra como medida preventiva de patógenos radiculares.", 0.20, "kg"),
    (7, "CAM-BH001-2025-002", "abono_verde",       "2025-05-01", "Semilla de marigold (Tagetes africana) sembrada en bordes de camas como planta compañera repelente de nematodos.", 0.10, "kg"),
]
for p in practicas:
    _row(ws, p[0] + 2, p)
_autowidth(ws)

# ── 18. COSECHAS (marketplace) ────────────────────────────────────────────────
ws = wb.create_sheet("Cosechas")
_hdr(ws, 1, ["#", "biohuerto", "cultivo_campaña", "categoria", "nombre_producto",
              "cantidad", "unidad", "precio_S/", "fecha_cosecha",
              "contacto_whatsapp", "estado"])
cosechas = [
    (1, "Biohuerto USAT – Bloque Norte", "CAM-BH001-2025-001", "Hortalizas de hoja", "Lechuga Batavia fresca",
     78.00, "kg",      3.50, "2025-04-08", "+51 950 123 456", "agotado"),
    (2, "Biohuerto USAT – Bloque Norte", "CAM-BH001-2025-002", "Bulbos y tubérculos", "Cebolla Roja Arequipeña",
     40.00, "kg",      2.80, "2025-07-20", "+51 950 123 456", "disponible"),
]
for c in cosechas:
    _row(ws, c[0] + 2, c)
_autowidth(ws)

# ── 19. PEDIDOS ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Pedidos")
_hdr(ws, 1, ["#", "comprador_username", "cosecha_producto", "cantidad", "precio_unitario_S/",
              "subtotal_S/", "total_pedido_S/", "estado", "notas", "mp_preference_id"])
pedidos = [
    (1, "maria.quispe",   "Lechuga Batavia fresca", 5.00,  3.50, 17.50, 17.50, "entregado",  "Entrega en portería USAT el 09/04.",  ""),
    (2, "juan.reyes",     "Lechuga Batavia fresca", 10.00, 3.50, 35.00, 35.00, "entregado",  "Para ensaladas del restaurante El Huerto.", "MP-123456"),
    (3, "carla.mendoza",  "Lechuga Batavia fresca", 8.00,  3.50, 28.00, 28.00, "entregado",  "",                                    ""),
    (4, "roberto.silva",  "Cebolla Roja Arequipeña",5.00,  2.80, 14.00, 14.00, "pendiente",  "Confirmar disponibilidad de pesaje.",  ""),
]
for p in pedidos:
    _row(ws, p[0] + 2, p)
_autowidth(ws)

# ── guardar ───────────────────────────────────────────────────────────────────
import os
out_dir = os.path.join(os.path.dirname(__file__))
out_path = os.path.join(out_dir, "biohuerto_datos_ejemplo.xlsx")
wb.save(out_path)
print(f"OK  Archivo generado: {out_path}")
